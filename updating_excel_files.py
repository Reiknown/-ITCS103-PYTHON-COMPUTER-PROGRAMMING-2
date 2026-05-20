import tkinter as tk
import openpyxl as op
from tkinter import ttk, messagebox


def display():
    
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("",tk.END, values=row)

def input_validation():
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if not first or not middle or not last or not birth:
        messagebox.showerror("Error", "ALL fields required!")
        return False
    
    if not birth.isdigit():
        messagebox.showerror("Errow", "Birthyear must be a number")
        return False
    
    return True

def saving():
    if not input_validation():
        return
    
    fname = fname_entry.get()
    mname = mname_entry.get()
    lname = lname_entry.get()
    by = int(birth_entry.get())

    age = 2026 - by

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id, lname, fname, mname, by, age])
    workbook.save("excelDB.xlsx")

    messagebox.showinfo("Success", "Record added successfully")
    display()

def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, 'values')

    if values:
        fname_entry.delete(0, tk.END)
        mname_entry.delete(0, tk.END)
        lname_entry.delete(0, tk.END)
        birth_entry.delete(0, tk.END)

        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        lname_entry.insert(0, values[1])
        birth_entry.insert(0, values[4])

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    if not input_validation():
        return
    
    values = table.item(selected, "values")
    record_id = values[0]

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())

    age = 2026 - birth

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for rows in sheet.iter_rows(min_row=2):
        if str(rows[0].value) == str(record_id):
            rows[1].value = last
            rows[2].value = first
            rows[3].value = middle
            rows[4].value = birth
            rows[5].value = age
    workbook.save("excelDB.xlsx")
    messagebox.showinfo("success", "Record updated successfully")
    display()

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return
    
    values = table.item(selected, "values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm", "Ar you sure you want to delete this record")
    if not confirm:
        return
    
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break
    
    workbook.save("excelDB.xlsx")
    messagebox.showinfo("Success", "Record deleted succesfully")
    display()


window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=1,columnspan=2)

update_btn = tk.Button(window, text="Update", command=update)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink", command=saving)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white", command=delete)
delete_btn.grid(row=6, column=3)

table = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for headings in ("ID","Last","First","Middle","BirthYear","Age"):
    table.heading(headings, text=headings)
    table.column(headings, anchor="center")
table.grid(row=7, column=0, columnspan=4)

table.bind("<<TreeviewSelect>>", auto_populate)

display()
window.mainloop()
